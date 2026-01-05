# SmartQSM - project-lightlin.github.io
# 
# Copyright (C) 2025-, YANG Jie <nj_yang_jie@foxmail.com>
# All rights reserved.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or 
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU Affero General Public License for more details.
# 
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

import open3d.visualization.gui as gui
import os
import yaml
import threading
from tkinter.filedialog import askopenfilename, asksaveasfilename
from _common_import import *
import openpyxl
from core.stand_structure import StandStructure
import traceback
import pandas as pd

APP_DIR = os.path.dirname(os.path.abspath(__file__))

class App:
    def __init__(self):
        self._window = gui.Application.instance.create_window(
            "Stand Structurer",
            640,
            720
        )

        self._main_layout = gui.Vert(
            5,
            gui.Margins(5, 5, 5, 5)
        )

        data_import_layout = gui.Vert()
        self._main_layout.add_child(data_import_layout)
        
        file_path_layout = gui.Horiz()
        data_import_layout.add_child(file_path_layout)
        
        file_path_layout.add_child( gui.Label("Stand data:"))
        
        self._opened_file_text_edit = gui.TextEdit()
        file_path_layout.add_child(self._opened_file_text_edit)
        self._opened_file_text_edit.enabled = False

        button_open_file = gui.Button("Open File")
        data_import_layout.add_child(button_open_file)
        button_open_file.set_on_clicked(self._on_button_open_file_clicked)

        feature_mapping_layout = gui.Vert()
        self._main_layout.add_child(feature_mapping_layout)

        feature_mapping_layout.add_child(gui.Label("Map features:"))
        features_tree_view = gui.TreeView()
        feature_mapping_layout.add_child(features_tree_view)
        
        item_name_id = features_tree_view.add_item(features_tree_view.get_root_item(), gui.Label("Name:"))
        self._name_combo_box = gui.Combobox()
        features_tree_view.add_item(item_name_id, self._name_combo_box)

        item_species_id = features_tree_view.add_item(features_tree_view.get_root_item(), gui.Label("Species:"))
        self._species_combo_box = gui.Combobox()
        features_tree_view.add_item(item_species_id, self._species_combo_box)

        item_x_id = features_tree_view.add_item(features_tree_view.get_root_item(), gui.Label("X (m):"))
        self._x_combo_box = gui.Combobox()
        features_tree_view.add_item(item_x_id, self._x_combo_box)

        item_y_id = features_tree_view.add_item(features_tree_view.get_root_item(), gui.Label("Y (m):"))
        self._y_combo_box = gui.Combobox()
        features_tree_view.add_item(item_y_id, self._y_combo_box)

        item_tree_height_id = features_tree_view.add_item(features_tree_view.get_root_item(), gui.Label("Tree height (m):"))
        self._tree_height_combo_box = gui.Combobox()
        features_tree_view.add_item(item_tree_height_id, self._tree_height_combo_box)

        item_dbh_id = features_tree_view.add_item(features_tree_view.get_root_item(), gui.Label("DBH (cm):"))
        self._dbh_combo_box = gui.Combobox()
        features_tree_view.add_item(item_dbh_id, self._dbh_combo_box)

        item_crown_width_id = features_tree_view.add_item(features_tree_view.get_root_item(), gui.Label("Crown width (m):"))
        self._crown_width_combo_box = gui.Combobox()
        features_tree_view.add_item(item_crown_width_id, self._crown_width_combo_box)

        language_setting_layout = gui.Vert()
        self._main_layout.add_child(language_setting_layout)

        language_setting_layout.add_child(gui.Label("Output language:"))
        self._language_list_view = gui.ListView()
        language_setting_layout.add_child(self.
        _language_list_view)
        self._language_list_view.set_max_visible_items(3)

        button_export = gui.Button("Export")
        self._main_layout.add_child(button_export)
        button_export.set_on_clicked(self._on_button_export_clicked)

        self._window.add_child(
            self._main_layout
        )
        self._window.set_on_layout(self._on_layout)
        self._window.set_on_close(self._on_close)

        # Initialize
        self._thread = None
        self._initialize_translations()

        return
    
    def _on_close(self):
        if self._thread is not None and self._thread.is_alive():
            self._thread._stop()
        gui.Application.instance.quit()
    
    def _on_layout(self, layout_content):
        content_rect = self._window.content_rect
        self._main_layout.frame = gui.Rect(
            content_rect.x,
            content_rect.y,
            content_rect.width,
            content_rect.height
        )
        return
    
    def _on_button_open_file_clicked(self):
        if self._thread is not None and self._thread.is_alive():
            return
        def select_and_load_file():
            file_path = askopenfilename(
                title="Open",
                filetypes=[("Excel workbook", "*.xlsx")]
            )
            if file_path:
                gui.Application.instance.post_to_main_thread(self._window, lambda: self._load_excel(file_path))
            else:
                gui.Application.instance.post_to_main_thread(self._window, self._window.close_dialog)
            return
        show_snackbar(self._window, "Waiting...")
        self._thread = threading.Thread(target=select_and_load_file)
        self._thread.start()
        return

    def _load_excel(self, file_path):
        self._opened_file_text_edit.text_value = ""
        self._name_combo_box.clear_items()
        self._species_combo_box.clear_items()
        self._x_combo_box.clear_items()
        self._y_combo_box.clear_items()
        self._tree_height_combo_box.clear_items()
        self._dbh_combo_box.clear_items()
        self._crown_width_combo_box.clear_items()

        try:
            self._workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            self._window.show_message_box(
                "",
                "Invalid XLSX file."
            )
            return
        
        self._opened_file_text_edit.text_value = file_path
        sheetnames = self._workbook.sheetnames

        dialog = gui.Dialog("")

        dialog_layout = gui.Vert(5, gui.Margins(5, 5, 5, 5))
        dialog.add_child(dialog_layout)

        dialog_layout.add_child(gui.Label("Select target sheet:"))
        self._sheetname_list_view = gui.ListView()
        dialog_layout.add_child(self._sheetname_list_view)
        self._sheetname_list_view.set_items(sheetnames)
        self._sheetname_list_view.selected_index = 0
        self._sheetname_list_view.set_max_visible_items(5)

        button_confirm = gui.Button("Confirm")
        dialog_layout.add_child(button_confirm)
        button_confirm.set_on_clicked(self._on_button_confirm_clicked)
        
        self._window.show_dialog(dialog)
        return
    
    def _on_button_confirm_clicked(self):
        self._window.close_dialog()

        sheetname = self._sheetname_list_view.selected_value
        sheet = self._workbook[sheetname]

        self._id_to_values = {}
        for column, column_cells in enumerate(sheet.iter_cols()):
            column_values = [cell.value for cell in list(column_cells)]

            self._name_combo_box.add_item(str(column_values[0]))
            self._species_combo_box.add_item(str(column_values[0]))
            self._x_combo_box.add_item(str(column_values[0]))
            self._y_combo_box.add_item(str(column_values[0]))
            self._tree_height_combo_box.add_item(str(column_values[0]))
            self._dbh_combo_box.add_item(str(column_values[0]))
            self._crown_width_combo_box.add_item(str(column_values[0]))

            self._id_to_values[column] = column_values[1:]

        self._workbook = None
        return
    
    def _on_button_export_clicked(self):

        column_name_to_values = {}
        language = self._language_list_view.selected_value
        try:
            names = self._id_to_values[self._name_combo_box.selected_index]
            column_name_to_values[self._translate("name", language)] = names
            stand_structure = StandStructure(
                names=names,
                species=self._id_to_values[self._species_combo_box.selected_index],
                x=np.array(self._id_to_values[self._x_combo_box.selected_index], dtype=np.float64),
                y=np.array(self._id_to_values[self._y_combo_box.selected_index], dtype=np.float64),
                heights=np.array(self._id_to_values[self._tree_height_combo_box.selected_index], dtype=np.float32),
                dbhs=np.array(self._id_to_values[self._dbh_combo_box.selected_index], dtype=np.float32),
                crown_widths=np.array(self._id_to_values[self._crown_width_combo_box.selected_index], dtype=np.float32)
            )

            column_name_to_values[self._translate("uniform_angle_index", language)] = stand_structure.uniform_angle_indices
            column_name_to_values[self._translate("hegyi_competition_index", language)] = stand_structure.hegyi_competition_indices
            column_name_to_values[self._translate("mingling", language)] = stand_structure.minglings
            column_name_to_values[self._translate("tree_species_diversity_mingling", language)] = stand_structure.tree_species_diversity_minglings
            column_name_to_values[self._translate("diameter_dominance", language)] = stand_structure.diameter_dominances
            column_name_to_values[self._translate("crowdedness", language)] = stand_structure.crowdednesses
            column_name_to_values[self._translate("openness", language)] = stand_structure.opennesses
            column_name_to_values[self._translate("within_unit_species_richness", language)] = stand_structure.within_unit_species_richnesses
            
            for i in range(stand_structure.unit_size):
                column_name_to_values[self._translate(f"neighbor_{i + 1}", language)] = [neighbor_names[i] for neighbor_names in stand_structure.neighbor_names_per_point]
                column_name_to_values[self._translate(f"distance_to_neighbor_{i + 1}_m", language)] = [distances_to_neighbors[i] for distances_to_neighbors in stand_structure.distances_to_neighbors_per_point]
            
            output_data_frame = pd.DataFrame(column_name_to_values)
            
        except Exception:
            self._window.show_message_box(
                "",
                "Encountered error:\n" + traceback.format_exc()
            )
            return
        
        def save_output_data_frame(output_data_frame):
            file_path = asksaveasfilename(
                filetypes=[("Excel files", "*.xlsx")],
                title="Save"
            )
            

            if not file_path:
                gui.Application.instance.post_to_main_thread(self._window, lambda: self._window.show_message_box(
                    "",
                    "Save canceled."
                ))
                return
            
            if not file_path.endswith(".xlsx"):
                file_path += ".xlsx"
            try:
                output_data_frame.to_excel(file_path, index=False)
                gui.Application.instance.post_to_main_thread(self._window, lambda: self._window.show_message_box(
                    "",
                    "Saved!"
                ))
            except Exception:
                gui.Application.instance.post_to_main_thread(self._window, lambda: self._window.show_message_box(
                    "",
                    "Encountered error:\n" + traceback.format_exc()
                ))
            return
        show_snackbar(self._window, "Waiting...")
        threading.Thread(target=lambda: save_output_data_frame(output_data_frame)).start()
        return

    def _translate(self, key, language):
        text = key
        try:
            text = self._language_to_dict[language][key]
        except Exception:
            pass
        return text
    
    def _initialize_translations(self):
        self._language_to_dict = {}
        try:
            with open(os.path.join(APP_DIR, "stand_structurer_translations.yaml"), "r", encoding="utf-8") as f:
                self._language_to_dict.update(yaml.safe_load(f))
        except Exception:
            pass
        self._language_list_view.set_items([""]+list(self._language_to_dict.keys()))
        return

def main():
    gui.Application.instance.initialize()
    set_default_font_description()
    App()
    gui.Application.instance.run()

if __name__ == "__main__":
    check_update()
    main()